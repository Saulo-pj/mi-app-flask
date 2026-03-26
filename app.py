from __future__ import annotations

import os
import json
from datetime import datetime, date, time, timedelta
from io import BytesIO
import re
from pathlib import Path
from uuid import uuid4
from collections import defaultdict

from flask import Flask, jsonify, redirect, render_template, request, session, send_file, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, select, delete, func
from sqlalchemy.orm import Session

from models import (
    Ajuste,
    Area,
    Categoria,
    ChecklistHistorico,
    ChecklistPedido,
    ChecklistProductoOculto,
    DetallePedido,
    InventarioSede,
    Producto,
    Sede,
    Turno,
    Unidad,
    CierreCaja,
    Movimiento,
    UsuarioGestion,
    get_engine,
    init_db,
    seed_data,
)


app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "dev-secret")

database_url = os.getenv("DATABASE_URL")
if database_url:
    database_url = database_url.replace("postgres://", "postgresql://")
else:
    database_url = "sqlite:///database.db"
app.config["SQLALCHEMY_DATABASE_URI"] = database_url

RESTAURANT_NAME = "Punto 29"
CENTRAL_SEDE_ID = 1
SUBAREAS_CONFIG: dict[str, list[dict[str, str]]] = {
    "COC": [
        {"nombre": "Cocina caliente", "badge": "bg-danger"},
        {"nombre": "Cocina fría", "badge": "bg-info"},
        {"nombre": "Lavadero", "badge": "bg-secondary"},
        {"nombre": "Mise en place", "badge": "bg-warning text-dark"},
    ],
    "SAL": [
        {"nombre": "Salón", "badge": "bg-primary"},
    ],
}
SUBAREA_TO_AREA: dict[str, str] = {}
for _area_id, items in SUBAREAS_CONFIG.items():
    for item in items:
        key = str(item.get("nombre", "")).strip().lower()
        if key:
            SUBAREA_TO_AREA[key] = _area_id
AREA_ORDER = {area_id: idx for idx, area_id in enumerate(SUBAREAS_CONFIG.keys())}
SUBAREA_POSITION: dict[str, dict[str, int]] = {
    area_id: {
        str(item.get("nombre", "")).strip().lower(): idx
        for idx, item in enumerate(items)
        if item.get("nombre")
    }
    for area_id, items in SUBAREAS_CONFIG.items()
}

engine = get_engine(database_url)

CREATION_PASSWORD_KEY = "creation_password"
DEFAULT_CREATION_PASSWORD = os.getenv("CHECKLIST_CREATION_PASSWORD", "almacen123")
DEFAULT_CHECKLIST_STOCK = float(os.getenv("CHECKLIST_STOCK_ON_CREATE", "20"))
DEFAULT_ALMACEN_CREDENTIALS = {"usuario": "71840465", "password": "71840465"}
DEFAULT_USER_MANAGEMENT_COCINA = [
    {"label": "Sede 17 · Día (Cocina)", "username": "sede17diacocina", "password": "2929", "id_sede": 17, "id_turno": "DIA", "id_area": "COC"},
    {"label": "Sede 17 · Noche (Cocina)", "username": "sede17nococina", "password": "2929", "id_sede": 17, "id_turno": "NOC", "id_area": "COC"},
    {"label": "Sede 20 · Día (Cocina)", "username": "sede20diacocina", "password": "2929", "id_sede": 20, "id_turno": "DIA", "id_area": "COC"},
    {"label": "Sede 20 · Noche (Cocina)", "username": "sede20nococina", "password": "2929", "id_sede": 20, "id_turno": "NOC", "id_area": "COC"},
]
DEFAULT_USER_MANAGEMENT_CAJA = [
    {"label": "Sede 17 · Día (Caja)", "username": "caja17dia", "password": "2929", "id_sede": 17, "id_turno": "DIA", "id_area": "SAL"},
    {"label": "Sede 17 · Noche (Caja)", "username": "caja17noche", "password": "2929", "id_sede": 17, "id_turno": "NOC", "id_area": "SAL"},
    {"label": "Sede 20 · Día (Caja)", "username": "caja20dia", "password": "2929", "id_sede": 20, "id_turno": "DIA", "id_area": "SAL"},
    {"label": "Sede 20 · Noche (Caja)", "username": "caja20noche", "password": "2929", "id_sede": 20, "id_turno": "NOC", "id_area": "SAL"},
]
DEFAULT_USER_MANAGEMENT_CIERRE = [
    {"label": "Sede 17 · Día (Cierre 1)", "username": "cierre17dia1", "password": "2929", "id_sede": 17, "id_turno": "DIA", "id_area": "SAL"},
    {"label": "Sede 17 · Día (Cierre 2)", "username": "cierre17dia2", "password": "2929", "id_sede": 17, "id_turno": "DIA", "id_area": "SAL"},
    {"label": "Sede 17 · Día (Cierre 3)", "username": "cierre17dia3", "password": "2929", "id_sede": 17, "id_turno": "DIA", "id_area": "SAL"},
    {"label": "Sede 17 · Día (Cierre 4)", "username": "cierre17dia4", "password": "2929", "id_sede": 17, "id_turno": "DIA", "id_area": "SAL"},
]
DEFAULT_USER_MANAGEMENT_MOVIMIENTOS = [
    {"label": "Personal Almacén 1", "username": "almacen1", "password": "2929", "id_sede": CENTRAL_SEDE_ID, "id_turno": "DIA", "id_area": "COC"},
    {"label": "Personal Almacén 2", "username": "almacen2", "password": "2929", "id_sede": CENTRAL_SEDE_ID, "id_turno": "DIA", "id_area": "COC"},
    {"label": "Personal Almacén 3", "username": "almacen3", "password": "2929", "id_sede": CENTRAL_SEDE_ID, "id_turno": "DIA", "id_area": "COC"},
    {"label": "Personal Almacén 4", "username": "almacen4", "password": "2929", "id_sede": CENTRAL_SEDE_ID, "id_turno": "DIA", "id_area": "COC"},
]
DEFAULT_USER_MANAGEMENT = {
    "cocina": DEFAULT_USER_MANAGEMENT_COCINA,
    "caja": DEFAULT_USER_MANAGEMENT_CAJA,
    "cierre": DEFAULT_USER_MANAGEMENT_CIERRE,
    "movimientos": DEFAULT_USER_MANAGEMENT_MOVIMIENTOS,
}

DEFAULT_ADMIN_CREDENTIALS = {"usuario": "admin", "password": "admin"}
ADMIN_USER_KEY = "admin_usuario"
ADMIN_PASSWORD_KEY = "admin_password"
ALMACEN_USER_KEY = "almacen_usuario"
ALMACEN_PASSWORD_KEY = "almacen_password"


def get_setting_value(session: Session, key: str, default: str) -> str:
    ajuste = session.get(Ajuste, key)
    return ajuste.Valor if ajuste else default


def set_setting_value(session: Session, key: str, value: str) -> None:
    ajuste = session.get(Ajuste, key)
    if ajuste:
        ajuste.Valor = value
    else:
        session.add(Ajuste(Clave=key, Valor=value))

MAX_CONCURRENT_SESSIONS = int(os.getenv("MAX_CONCURRENT_SESSIONS", "5"))
SESSION_ACTIVITY_WINDOW = timedelta(minutes=30)
ACTIVE_SESSION_TOKENS: dict[str, datetime] = {}


UserEntry = dict[str, str | int | None]


def load_user_management_state() -> dict[str, list[UserEntry]]:
    grouped: dict[str, list[UserEntry]] = defaultdict(list)
    with Session(engine) as db:
        rows = db.scalars(select(UsuarioGestion).order_by(UsuarioGestion.Perfil, UsuarioGestion.Orden)).all()
        for row in rows:
            grouped.setdefault(row.Perfil, []).append(
                {
                    "label": row.Label,
                    "username": row.Username,
                    "password": row.Password,
                    "id_sede": row.ID_Sede,
                    "id_turno": row.ID_Turno or "",
                    "id_area": row.ID_Area or "",
                }
            )
    return grouped


def persist_user_group(perfil: str, entries: list[UserEntry]) -> None:
    with Session(engine) as db:
        db.execute(delete(UsuarioGestion).where(UsuarioGestion.Perfil == perfil))
        for idx, entry in enumerate(entries):
            db.add(
                UsuarioGestion(
                    Perfil=perfil,
                    Label=entry.get("label", f"Usuario {idx + 1}"),
                    Username=entry.get("username", ""),
                    Password=entry.get("password", ""),
                    ID_Sede=entry.get("id_sede"),
                    ID_Turno=entry.get("id_turno") or "",
                    ID_Area=entry.get("id_area") or "",
                    Orden=idx,
                )
            )
        db.commit()


def ensure_default_user_management(session: Session) -> None:
    existing = session.scalar(select(func.count()).select_from(UsuarioGestion)) or 0
    if existing:
        return
    for perfil, defaults in DEFAULT_USER_MANAGEMENT.items():
        for idx, default in enumerate(defaults):
            session.add(
                UsuarioGestion(
                    Perfil=perfil,
                    Label=default["label"],
                    Username=default["username"],
                    Password=default["password"],
                    ID_Sede=default.get("id_sede"),
                    ID_Turno=default.get("id_turno", ""),
                    ID_Area=default.get("id_area", ""),
                    Orden=idx,
                )
            )
    session.commit()


init_db(engine)
with Session(engine) as db:
    seed_data(db)
    ensure_default_user_management(db)


def _parse_user_updates(prefix: str, reference_list: list[UserEntry], form_data) -> tuple[list[UserEntry] | None, str | None]:
    updated: list[UserEntry] = []
    for idx, reference in enumerate(reference_list):
        label_key = f"{prefix}_label_{idx}"
        label = str(form_data.get(label_key, reference.get("label", f"Usuario {idx + 1}"))).strip() or reference.get("label", f"Usuario {idx + 1}")
        username = str(form_data.get(f"{prefix}_username_{idx}", "")).strip()
        password = str(form_data.get(f"{prefix}_password_{idx}", "")).strip()
        if not username or not password:
            return None, f"Completa usuario y contraseña para {label}"
        sede_raw = form_data.get(f"{prefix}_sede_{idx}")
        if sede_raw in (None, ""):
            sede_raw = reference.get("id_sede")
        try:
            id_sede = int(sede_raw)
        except (TypeError, ValueError):
            id_sede = None
        if id_sede is None:
            return None, f"Selecciona una sede válida para {label}"
        turno_raw = form_data.get(f"{prefix}_turno_{idx}")
        if turno_raw in (None, ""):
            turno_raw = reference.get("id_turno")
        id_turno = str(turno_raw or "").strip()
        if not id_turno:
            return None, f"Selecciona un turno para {label}"
        area_raw = form_data.get(f"{prefix}_area_{idx}")
        if area_raw in (None, ""):
            area_raw = reference.get("id_area")
        id_area = str(area_raw or "").strip()
        if not id_area:
            return None, f"Selecciona un área para {label}"
        updated.append({
            "label": label,
            "username": username,
            "password": password,
            "id_sede": id_sede,
            "id_turno": id_turno,
            "id_area": id_area,
        })
    return updated, None


def _cleanup_inactive_sessions() -> None:
    now = datetime.utcnow()
    stale_tokens = [token for token, last_seen in ACTIVE_SESSION_TOKENS.items() if now - last_seen > SESSION_ACTIVITY_WINDOW]
    for token in stale_tokens:
        ACTIVE_SESSION_TOKENS.pop(token, None)


def _is_session_slot_available() -> bool:
    _cleanup_inactive_sessions()
    token = session.get("concurrency_token")
    if token and token in ACTIVE_SESSION_TOKENS:
        return True
    return len(ACTIVE_SESSION_TOKENS) < MAX_CONCURRENT_SESSIONS


def _reserve_session_slot() -> str:
    now = datetime.utcnow()
    token = session.get("concurrency_token")
    if token and token in ACTIVE_SESSION_TOKENS:
        ACTIVE_SESSION_TOKENS[token] = now
        return token
    new_token = str(uuid4())
    session["concurrency_token"] = new_token
    ACTIVE_SESSION_TOKENS[new_token] = now
    return new_token


def _release_session_slot() -> None:
    token = session.get("concurrency_token")
    if token:
        ACTIVE_SESSION_TOKENS.pop(token, None)


def find_user_entry(username: str, password: str, groups: str | list[str] | None = None) -> UserEntry | None:
    username = username.strip()
    password = password.strip()
    if not username or not password:
        return None
    if groups is None:
        group_names: list[str] | None = None
    elif isinstance(groups, str):
        group_names = [groups]
    else:
        group_names = groups
    consulta = select(UsuarioGestion)
    if group_names:
        consulta = consulta.where(UsuarioGestion.Perfil.in_(group_names))
    with Session(engine) as db:
        for row in db.scalars(consulta).all():
            if row.Username == username and row.Password == password:
                return {
                    "label": row.Label,
                    "username": row.Username,
                    "password": row.Password,
                    "id_sede": row.ID_Sede,
                    "id_turno": row.ID_Turno or "",
                    "id_area": row.ID_Area or "",
                }
    return None


def _is_sede_or_admin_profile() -> bool:
    return session.get("perfil") in ("sede", "admin")


def _is_almacen_or_admin_profile() -> bool:
    return session.get("perfil") in ("almacen", "admin")


def actualizar_estado_general_pedido(pedido: ChecklistPedido) -> None:
    if not pedido:
        return
    detalles = pedido.detalles or []
    if not detalles:
        pedido.Estado_General = "Pendiente"
        return
    if all(det.Estado_Sede == "Recibido" for det in detalles):
        pedido.Estado_General = "Recibido"
        return
    if any(det.Estado_Sede == "Enviado" for det in detalles):
        pedido.Estado_General = "Parcial"
        return
    pedido.Estado_General = "Pendiente"


def pedido_pertenece_al_contexto(pedido: ChecklistPedido, ctx: dict[str, int | str]) -> bool:
    if not pedido:
        return False
    return (
        pedido.ID_Sede == ctx.get("id_sede")
        and pedido.ID_Turno == ctx.get("id_turno")
        and pedido.ID_Area == ctx.get("id_area")
    )


def subarea_badge(area_id: str | None, subarea: str | None) -> str:
    if not area_id or not subarea:
        return "bg-secondary"
    normalized = subarea.strip().lower()
    for item in SUBAREAS_CONFIG.get(area_id, []):
        nombre = str(item.get("nombre", "")).strip().lower()
        if nombre == normalized:
            return item.get("badge", "bg-secondary")
    return "bg-secondary"


def normalize_subarea_name(subarea: str | None) -> str:
    return str(subarea or "").strip().lower()


_ID_SUFFIX_PATTERN = re.compile(r"(\d+)$")


def get_or_create_creation_password(db: Session) -> Ajuste:
    ajuste = db.get(Ajuste, CREATION_PASSWORD_KEY)
    if not ajuste:
        ajuste = Ajuste(Clave=CREATION_PASSWORD_KEY, Valor=generate_password_hash(DEFAULT_CREATION_PASSWORD))
        db.add(ajuste)
        db.commit()
        db.refresh(ajuste)
    return ajuste


def verify_creation_password(db: Session, password: str | None) -> bool:
    ajuste = get_or_create_creation_password(db)
    return check_password_hash(ajuste.Valor, (password or "").strip())


def update_creation_password(db: Session, new_password: str) -> None:
    ajuste = get_or_create_creation_password(db)
    ajuste.Valor = generate_password_hash(new_password.strip())
    db.commit()


def next_product_id(db: Session) -> str:
    productos = db.scalars(select(Producto.ID_Producto)).all()
    max_number = 0
    for producto_id in productos:
        if not producto_id:
            continue
        match = _ID_SUFFIX_PATTERN.search(producto_id)
        if not match:
            continue
        try:
            max_number = max(max_number, int(match.group(1)))
        except ValueError:
            continue
    next_number = max_number + 1
    return f"PROD{next_number:04d}"


def crear_producto_en_inventario(
    db: Session,
    id_producto: str,
    nombre: str,
    id_area: str,
    subarea: str,
    id_categoria: int,
    id_unidad: int,
    stock_central: float,
    punto_minimo: float = 5.0,
    extra_sedes: list[int] | None = None,
):
    producto = Producto(
        ID_Producto=id_producto,
        Nombre_Producto=nombre,
        ID_Area=id_area,
        Subarea=subarea,
        ID_Categoria=id_categoria,
        ID_Unidad=id_unidad,
        Estado="Activo",
    )
    db.add(producto)
    db.flush()

    inventario = db.get(InventarioSede, (CENTRAL_SEDE_ID, id_producto))
    if not inventario:
        inventario = InventarioSede(
            ID_Sede=CENTRAL_SEDE_ID,
            ID_Producto=id_producto,
            Stock_Actual=stock_central,
            Punto_Minimo=punto_minimo,
        )
        db.add(inventario)
    else:
        inventario.Stock_Actual = stock_central
        inventario.Punto_Minimo = punto_minimo

    if extra_sedes:
        for sede_id in extra_sedes:
            if not sede_id or sede_id == CENTRAL_SEDE_ID:
                continue
            extra_inv = db.get(InventarioSede, (sede_id, id_producto))
            if not extra_inv:
                extra_inv = InventarioSede(
                    ID_Sede=sede_id,
                    ID_Producto=id_producto,
                    Stock_Actual=0,
                    Punto_Minimo=0,
                )
                db.add(extra_inv)
            else:
                if extra_inv.Stock_Actual is None:
                    extra_inv.Stock_Actual = 0
                if extra_inv.Punto_Minimo is None:
                    extra_inv.Punto_Minimo = 0

    return producto


def get_or_create_pedido(db: Session, ctx: dict[str, int | str], fecha_filtro: datetime | None = None) -> ChecklistPedido:
    filtros = [
        ChecklistPedido.ID_Sede == ctx["id_sede"],
        ChecklistPedido.ID_Turno == ctx["id_turno"],
        ChecklistPedido.ID_Area == ctx["id_area"],
    ]
    consulta = select(ChecklistPedido).where(*filtros)
    if fecha_filtro:
        inicio = datetime.combine(fecha_filtro.date(), time.min)
        fin = inicio + timedelta(days=1)
        consulta = consulta.where(
            ChecklistPedido.Fecha >= inicio,
            ChecklistPedido.Fecha < fin,
        )
    consulta = consulta.order_by(ChecklistPedido.Fecha.desc())

    pedido = db.scalars(consulta).first()
    if pedido:
        return pedido

    pedido = ChecklistPedido(
        ID_Sede=ctx["id_sede"],
        ID_Turno=ctx["id_turno"],
        ID_Area=ctx["id_area"],
        ID_Usuario=ctx.get("id_usuario", "sin_usuario"),
        Fecha=fecha_filtro or datetime.now(),
    )
    db.add(pedido)
    db.flush()
    db.commit()
    db.refresh(pedido)
    return pedido


def obtener_fecha_checklist() -> datetime:
    raw = session.get("checklist_fecha")
    if raw:
        try:
            parsed = datetime.fromisoformat(str(raw))
        except ValueError:
            parsed = None
        if parsed:
            return parsed
    ahora = datetime.now()
    return ahora


def actualizar_historial_checklist(
    db: Session,
    ctx: dict[str, int | str],
    fecha: datetime,
    detalles: list[DetallePedido],
) -> None:
    fecha_dia = fecha.date()
    db.query(ChecklistHistorico).filter(
        ChecklistHistorico.ID_Sede == ctx["id_sede"],
        ChecklistHistorico.ID_Area == ctx["id_area"],
        ChecklistHistorico.Fecha == fecha_dia,
    ).delete()
    if not detalles:
        return
    for detalle in detalles:
        producto = db.get(Producto, detalle.ID_Producto)
        nombre = producto.Nombre_Producto if producto else detalle.ID_Producto
        db.add(
            ChecklistHistorico(
                Fecha=fecha_dia,
                ID_Sede=ctx["id_sede"],
                ID_Area=ctx["id_area"],
                ID_Producto=detalle.ID_Producto,
                Nombre_Producto=nombre,
            )
        )


def obtener_lista_historica(db: Session, ctx: dict[str, int | str], fecha: datetime) -> list[str]:
    fecha_dia = fecha.date()
    stmt = select(ChecklistHistorico.Nombre_Producto).where(
        and_(
            ChecklistHistorico.ID_Sede == ctx["id_sede"],
            ChecklistHistorico.ID_Area == ctx["id_area"],
            ChecklistHistorico.Fecha == fecha_dia,
        )
    ).order_by(ChecklistHistorico.ID_Historico)
    return db.scalars(stmt).all()


def _safe_float(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _parse_gastos_detalle(raw: str | None) -> list[dict[str, float | str]]:
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except ValueError:
        return []
    if not isinstance(parsed, list):
        return []
    normalized: list[dict[str, float | str]] = []
    for item in parsed:
        if not isinstance(item, dict):
            continue
        descripcion = str(item.get("descripcion", "")).strip()
        monto = _safe_float(item.get("monto", 0))
        normalized.append({"descripcion": descripcion, "monto": monto})
    return normalized


def _build_closing_payload(cierre: CierreCaja | None) -> dict[str, float | str | list[dict[str, float | str]]]:
    if not cierre:
        return {
            "monto_inicial": 0.0,
            "pos": 0.0,
            "yape": 0.0,
            "plin": 0.0,
            "efectivo": 0.0,
            "venta_sistema": 0.0,
            "total_actual": 0.0,
            "diferencia": 0.0,
            "observaciones": "",
            "gastos": [],
        }
    gastos = _parse_gastos_detalle(cierre.Gastos_Detalle)
    return {
        "monto_inicial": float(cierre.Monto_Inicial or 0),
        "pos": float(cierre.POS or 0),
        "yape": float(cierre.Yape or 0),
        "plin": float(cierre.Plin or 0),
        "efectivo": float(cierre.Efectivo or 0),
        "venta_sistema": float(cierre.Venta_Sistema or 0),
        "total_actual": float(cierre.Total_Actual or 0),
        "diferencia": float(cierre.Diferencia or 0),
        "observaciones": cierre.Observaciones or "",
        "gastos": gastos,
    }


def _build_dashboard_payload(
    today_records: list[dict[str, object]],
    recent_records: list[dict[str, object]],
    sede_info: list[dict[str, int | str]],
    turno_info: list[dict[str, str]],
    selected_date: date,
    selected_sede: int | None,
    selected_turno: str | None,
) -> dict[str, object]:
    def summable(record: dict[str, float | str | int | datetime | date], key: str) -> float:
        return _safe_float(record.get(key))

    def income_for(record: dict[str, float | str | int | datetime | date]) -> float:
        return (
            summable(record, "pos")
            + summable(record, "yape")
            + summable(record, "plin")
            + summable(record, "efectivo")
        )

    totals: dict[str, float] = {"ingresos": 0.0, "gastos": 0.0, "diferencia": 0.0}
    payments: dict[str, float] = {"pos": 0.0, "yape": 0.0, "plin": 0.0, "efectivo": 0.0}
    grouping: dict[str, dict[str, float]] = {}
    for record in today_records:
        incomes = income_for(record)
        gastos = summable(record, "gastos")
        key = f"{record.get('sede_id')}:{record.get('turno_id')}"
        entry = grouping.setdefault(key, {"ingresos": 0.0, "gastos": 0.0, "net": 0.0})
        entry["ingresos"] += incomes
        entry["gastos"] += gastos
        entry["net"] += incomes - gastos
        totals["ingresos"] += incomes
        totals["gastos"] += gastos
        totals["diferencia"] += summable(record, "diferencia")
        payments["pos"] += summable(record, "pos")
        payments["yape"] += summable(record, "yape")
        payments["plin"] += summable(record, "plin")
        payments["efectivo"] += summable(record, "efectivo")

    table_rows: list[dict[str, object]] = []
    table_totals = {"ingresos": 0.0, "gastos": 0.0, "ventas_netas": 0.0}
    for sede in sede_info:
        for turno in turno_info:
            key = f"{sede.get('id')}:{turno.get('id')}"
            entry = grouping.get(key, {"ingresos": 0.0, "gastos": 0.0, "net": 0.0})
            ingresos = entry.get("ingresos", 0.0)
            gastos = entry.get("gastos", 0.0)
            net = entry.get("net", ingresos - gastos)
            table_totals["ingresos"] += ingresos
            table_totals["gastos"] += gastos
            table_totals["ventas_netas"] += net
            table_rows.append(
                {
                    "sede": sede.get("name", "Sede"),
                    "turno": turno.get("name", "Turno"),
                    "ingresos": round(ingresos, 2),
                    "gastos": round(gastos, 2),
                    "ventas_netas": round(net, 2),
                }
            )

    distribution = [
        {
            "label": f"{row['sede']} · {row['turno']}",
            "value": row["ingresos"],
        }
        for row in table_rows
    ]

    history_days: list[datetime.date] = []
    cursor = selected_date - timedelta(days=6)
    while cursor <= selected_date:
        history_days.append(cursor)
        cursor += timedelta(days=1)
    daily_map: dict[datetime.date, dict[str, float]] = {
        day: {"ingresos": 0.0, "gastos": 0.0} for day in history_days
    }
    for record in recent_records:
        fecha = record.get("fecha")
        if fecha in daily_map:
            daily_map[fecha]["ingresos"] += income_for(record)
            daily_map[fecha]["gastos"] += summable(record, "gastos")

    trend = {
        "labels": [day.strftime("%d/%m") for day in history_days],
        "ingresos": [daily_map[day]["ingresos"] for day in history_days],
        "gastos": [daily_map[day]["gastos"] for day in history_days],
    }

    selected_sede_label = next((item["name"] for item in sede_info if item.get("id") == selected_sede), "Todas las sedes")
    selected_turno_label = next((item["name"] for item in turno_info if item.get("id") == selected_turno), "Todos los turnos")

    return {
        "filters": {
            "selected_date": selected_date.strftime("%Y-%m-%d"),
            "selected_sede": selected_sede,
            "selected_turno": selected_turno,
            "selected_sede_label": selected_sede_label,
            "selected_turno_label": selected_turno_label,
        },
        "kpis": {
            "total_ingresos": round(totals["ingresos"], 2),
            "total_gastos": round(totals["gastos"], 2),
            "utilidad_neta": round(totals["ingresos"] - totals["gastos"], 2),
            "diferencia_vs_sistema": round(totals["diferencia"], 2),
        },
        "sedeOrder": sede_info,
        "turnoOrder": turno_info,
        "bySedeTurno": {
            key: {"value": round(value.get("ingresos", 0.0), 2), "net": round(value.get("net", 0.0), 2)}
            for key, value in grouping.items()
        },
        "table_rows": table_rows,
        "table_total": {
            "ingresos": round(table_totals["ingresos"], 2),
            "gastos": round(table_totals["gastos"], 2),
            "ventas_netas": round(table_totals["ventas_netas"], 2),
        },
        "payments": {key: round(value, 2) for key, value in payments.items()},
        "distribution": distribution,
        "incomes_vs_gastos": {"ingresos": round(totals["ingresos"], 2), "gastos": round(totals["gastos"], 2)},
        "trend": trend,
        "panel": {
            "pos": round(payments["pos"], 2),
            "yape": round(payments["yape"], 2),
            "plin": round(payments["plin"], 2),
            "efectivo": round(payments["efectivo"], 2),
            "total": round(totals["ingresos"], 2),
        },
    }

@app.get("/")
def index():
    return redirect(url_for("login"))


@app.before_request
def refresh_session_activity():
    _cleanup_inactive_sessions()
    token = session.get("concurrency_token")
    if token:
        ACTIVE_SESSION_TOKENS[token] = datetime.utcnow()


@app.route("/login", methods=["GET", "POST"])
def login():
    with Session(engine) as db:
        sedes = db.scalars(
            select(Sede).where(Sede.ID_Sede.in_([17, 20])).order_by(Sede.ID_Sede)
        ).all()
        areas = db.scalars(select(Area).order_by(Area.Nombre_Area)).all()
        turnos = db.scalars(select(Turno).order_by(Turno.Nombre_Turno)).all()
        admin_credentials = {
            "usuario": get_setting_value(db, ADMIN_USER_KEY, DEFAULT_ADMIN_CREDENTIALS["usuario"]),
            "password": get_setting_value(db, ADMIN_PASSWORD_KEY, DEFAULT_ADMIN_CREDENTIALS["password"]),
        }
        default_almacen_credentials = {
            "usuario": get_setting_value(db, ALMACEN_USER_KEY, DEFAULT_ALMACEN_CREDENTIALS["usuario"]),
            "password": get_setting_value(db, ALMACEN_PASSWORD_KEY, DEFAULT_ALMACEN_CREDENTIALS["password"]),
        }

    feedback = None
    feedback_type = None
    sticky_username = ""
    active_module = "almacen"
    concurrency_message = (
        f"Solo {MAX_CONCURRENT_SESSIONS} personas pueden usar la app al mismo tiempo. Intenta nuevamente en unos minutos."
    )
    if request.method == "POST":
        if not _is_session_slot_available():
            feedback = concurrency_message
            feedback_type = "warning"
        else:
            perfil = request.form.get("perfil", "sede")
            active_module = perfil
            if perfil == "almacen":
                username = request.form.get("almacen_usuario", "").strip()
                password = request.form.get("almacen_password", "")
                movimientos_entry = find_user_entry(username, password, "movimientos")
                if username == default_almacen_credentials["usuario"] and password == default_almacen_credentials["password"]:
                    session["perfil"] = perfil
                    session["movimientos_only"] = False
                    session["id_sede"] = CENTRAL_SEDE_ID
                    session["id_turno"] = ""
                    session["id_area"] = ""
                    session["id_usuario"] = username
                    _reserve_session_slot()
                    return redirect(url_for("almacen"))
                if movimientos_entry:
                    session["perfil"] = perfil
                    session["movimientos_only"] = True
                    session["id_sede"] = movimientos_entry.get("id_sede") or CENTRAL_SEDE_ID
                    session["id_turno"] = movimientos_entry.get("id_turno") or ""
                    session["id_area"] = movimientos_entry.get("id_area") or ""
                    session["id_usuario"] = movimientos_entry.get("username") or username
                    session["login_feedback"] = {
                        "type": "success",
                        "text": f"Inicio de sesión correcto ({movimientos_entry.get('label', movimientos_entry.get('username', 'usuario'))}).",
                    }
                    _reserve_session_slot()
                    return redirect(url_for("almacen"))
                feedback = "Usuario o contraseña inválidos"
                feedback_type = "danger"
                sticky_username = username
            else:
                session["movimientos_only"] = False
                if perfil == "sede":
                    username = request.form.get("sede_usuario", "").strip()
                    password = request.form.get("sede_password", "")
                    entry = find_user_entry(username, password, ["cocina", "caja"])
                elif perfil == "admin":
                    username = request.form.get("admin_usuario", "").strip()
                    password = request.form.get("admin_password", "")
                    valid_admin = username == admin_credentials["usuario"] and password == admin_credentials["password"]
                    if valid_admin:
                        session["perfil"] = "admin"
                        selected_sede = sedes[0].ID_Sede if sedes else None
                        selected_turno = turnos[0].ID_Turno if turnos else ""
                        selected_area = areas[0].ID_Area if areas else ""
                        session["id_sede"] = selected_sede
                        session["id_turno"] = selected_turno
                        session["id_area"] = selected_area
                        session["id_usuario"] = admin_credentials["usuario"]
                        session["login_feedback"] = {
                            "type": "success",
                            "text": "Inicio de sesión correcto (Administrador).",
                        }
                        _reserve_session_slot()
                        return redirect(url_for("checklist"))
                    feedback = "Usuario o contraseña inválidos"
                    feedback_type = "danger"
                    sticky_username = username
                    entry = None
                else:
                    username = request.form.get("cierre_usuario", "").strip()
                    password = request.form.get("cierre_password", "")
                    entry = find_user_entry(username, password, "cierre")
                if perfil != "admin" and not entry:
                    feedback = "Usuario o contraseña inválidos"
                    feedback_type = "danger"
                    sticky_username = username
                elif perfil != "admin":
                    session["perfil"] = perfil
                    session["id_sede"] = entry.get("id_sede")
                    session["id_turno"] = entry.get("id_turno")
                    session["id_area"] = entry.get("id_area")
                    session["id_usuario"] = entry.get("username") or username
                    session["login_feedback"] = {
                        "type": "success",
                        "text": f"Inicio de sesión correcto ({entry.get('label', entry.get('username', 'usuario'))}).",
                    }
                    _reserve_session_slot()
                    if perfil == "cierre":
                        sede_param = entry.get("id_sede")
                        turno_param = entry.get("id_turno")
                        return redirect(
                            url_for(
                                "cierre_caja",
                                id_sede=sede_param,
                                id_turno=turno_param,
                            )
                        )
                    return redirect(url_for("checklist"))

    return render_template(
        "login.html",
        sedes=sedes,
        areas=areas,
        turnos=turnos,
        restaurant_name=RESTAURANT_NAME,
        default_almacen_credentials=default_almacen_credentials,
        admin_credentials=admin_credentials,
        feedback=feedback,
        feedback_type=feedback_type,
        active_module=active_module,
        sticky_username=sticky_username,
        movimientos_users=load_user_management_state().get("movimientos", []),
    )


@app.route("/checklist", methods=["GET", "POST"])
def checklist():
    if not _is_sede_or_admin_profile():
        return redirect(url_for("login"))

    if request.method == "POST":
        if session.get("perfil") == "admin":
            errors: list[str] = []
            new_sede = request.form.get("admin_sede")
            new_turno = request.form.get("admin_turno")
            new_area = request.form.get("admin_area")
            contexto_actualizado: dict[str, int | str] = {}
            if new_sede:
                try:
                    contexto_actualizado["id_sede"] = int(new_sede)
                except ValueError:
                    errors.append("Selecciona una sede válida")
            else:
                errors.append("Selecciona una sede")
            if new_turno:
                contexto_actualizado["id_turno"] = new_turno
            else:
                errors.append("Selecciona un turno")
            if new_area:
                contexto_actualizado["id_area"] = new_area
            else:
                errors.append("Selecciona un área")
            if errors:
                session["admin_context_feedback"] = {
                    "type": "danger",
                    "text": errors[0],
                }
            else:
                session.update(contexto_actualizado)
                session["admin_context_feedback"] = {
                    "type": "success",
                    "text": "Contexto actualizado",
                }
        return redirect(url_for("checklist"))

    login_feedback = session.pop("login_feedback", None)
    admin_context_message = session.pop("admin_context_feedback", None)

    raw_fecha = request.args.get("fecha")
    fecha_filtro = None
    if raw_fecha:
        try:
            parsed = datetime.fromisoformat(raw_fecha)
        except ValueError:
            try:
                parsed = datetime.fromisoformat(f"{raw_fecha}T00:00:00")
            except ValueError:
                parsed = None
        if parsed:
            fecha_filtro = parsed
    if fecha_filtro is None:
        fecha_filtro = datetime.now()
    fecha_filtro = fecha_filtro.replace(hour=0, minute=0, second=0, microsecond=0)
    fecha_anterior = fecha_filtro - timedelta(days=1)
    session["checklist_fecha"] = fecha_filtro.isoformat()

    ctx = {
        "id_sede": session.get("id_sede"),
        "id_turno": session.get("id_turno"),
        "id_area": session.get("id_area"),
        "id_usuario": session.get("id_usuario"),
    }

    with Session(engine) as db:
        pedido = get_or_create_pedido(db, ctx, fecha_filtro)
        sede = db.get(Sede, ctx["id_sede"])
        area = db.get(Area, ctx["id_area"])
        turno = db.get(Turno, ctx["id_turno"])
        lista_historica = obtener_lista_historica(db, ctx, fecha_anterior)
        admin_sedes = []
        admin_turnos = []
        admin_areas = []
        if session.get("perfil") == "admin":
            admin_sedes = db.scalars(select(Sede).order_by(Sede.Nombre_Sede)).all()
            admin_turnos = db.scalars(select(Turno).order_by(Turno.Nombre_Turno)).all()
            admin_areas = db.scalars(select(Area).order_by(Area.Nombre_Area)).all()

    now = datetime.now()
    fecha_seleccionada = pedido.Fecha or fecha_filtro or now
    selected_date_label = fecha_seleccionada.strftime("%d/%m/%Y")
    selected_time_label = fecha_seleccionada.strftime("%H:%M")
    show_time = False
    fecha_input = fecha_seleccionada.strftime("%Y-%m-%d")

    is_editable = pedido.Estado_General in ("Pendiente", "Parcial")

    return render_template(
        "checklist.html",
        sede=sede,
        area=area,
        turno=turno,
        pedido_id=pedido.ID_Pedido,
        estado=pedido.Estado_General,
        editable=is_editable,
        restaurant_name=RESTAURANT_NAME,
        subareas_config=SUBAREAS_CONFIG,
        selected_date_label=selected_date_label,
        selected_time_label=selected_time_label,
        selected_datetime_input=fecha_input,
        show_time=show_time,
        lista_historica=lista_historica,
        fecha_anterior_label=fecha_anterior.strftime("%d/%m/%Y"),
        login_feedback=login_feedback,
        admin_context_message=admin_context_message,
        is_admin=session.get("perfil") == "admin",
        admin_sedes=admin_sedes,
        admin_turnos=admin_turnos,
        admin_areas=admin_areas,
    )


@app.get("/cierre-caja")
def cierre_caja():
    raw_sede = request.args.get("id_sede")
    raw_turno = request.args.get("id_turno")
    raw_fecha = request.args.get("fecha")
    selected_date = None
    if raw_fecha:
        try:
            selected_date = datetime.fromisoformat(raw_fecha).date()
        except ValueError:
            selected_date = None
    if not selected_date:
        selected_date = datetime.now().date()

    is_admin = session.get("perfil") == "admin"
    movimientos_only = bool(session.get("movimientos_only"))
    dashboard_today_records: list[dict[str, float | str | int | datetime | date]] = []
    dashboard_recent_records: list[dict[str, float | str | int | datetime | date]] = []
    sede_info: list[dict[str, int | str]] = []
    turno_info: list[dict[str, str]] = []
    recent_start = selected_date - timedelta(days=6)

    with Session(engine) as db:
        sedes = db.scalars(select(Sede).order_by(Sede.Nombre_Sede)).all()
        turnos = db.scalars(select(Turno).order_by(Turno.Nombre_Turno)).all()
        sede_info = [{"id": sede.ID_Sede, "name": sede.Nombre_Sede} for sede in sedes]
        turno_info = [{"id": turno.ID_Turno, "name": turno.Nombre_Turno} for turno in turnos]
        today_stmt = select(CierreCaja).where(CierreCaja.Fecha == selected_date)
        for cierre_row in db.scalars(today_stmt).all():
            dashboard_today_records.append({
                "fecha": cierre_row.Fecha,
                "sede_id": cierre_row.ID_Sede,
                "turno_id": cierre_row.ID_Turno,
                "pos": cierre_row.POS,
                "yape": cierre_row.Yape,
                "plin": cierre_row.Plin,
                "efectivo": cierre_row.Efectivo,
                "gastos": cierre_row.Total_Gastos,
                "diferencia": cierre_row.Diferencia,
            })
        recent_stmt = select(CierreCaja).where(
            CierreCaja.Fecha >= recent_start,
            CierreCaja.Fecha <= selected_date,
        ).order_by(CierreCaja.Fecha)
        for cierre_row in db.scalars(recent_stmt).all():
            dashboard_recent_records.append({
                "fecha": cierre_row.Fecha,
                "pos": cierre_row.POS,
                "yape": cierre_row.Yape,
                "plin": cierre_row.Plin,
                "efectivo": cierre_row.Efectivo,
                "gastos": cierre_row.Total_Gastos,
            })
        selected_sede = None
        if is_admin:
            if raw_sede:
                try:
                    selected_sede = int(raw_sede)
                except ValueError:
                    selected_sede = None
        else:
            selected_sede = session.get("id_sede")
        if selected_sede is None and sedes:
            selected_sede = sedes[0].ID_Sede
        if is_admin:
            selected_turno = raw_turno if raw_turno else (turnos[0].ID_Turno if turnos else "")
        else:
            selected_turno = session.get("id_turno") or (turnos[0].ID_Turno if turnos else "")
        cierre = None
        if selected_sede and selected_turno:
            cierre = db.scalar(
                select(CierreCaja).where(
                    CierreCaja.ID_Sede == selected_sede,
                    CierreCaja.ID_Turno == selected_turno,
                    CierreCaja.Fecha == selected_date,
                )
            )

    closing_data = _build_closing_payload(cierre)
    dashboard_data = None
    if is_admin:
        dashboard_data = _build_dashboard_payload(
            dashboard_today_records,
            dashboard_recent_records,
            sede_info,
            turno_info,
            selected_date,
            selected_sede,
            selected_turno,
        )
    sede_label = next((s.Nombre_Sede for s in sedes if s.ID_Sede == selected_sede), "Sede")
    turno_label = next((t.Nombre_Turno for t in turnos if t.ID_Turno == selected_turno), "Turno")
    return render_template(
        "cierre_caja.html",
        sedes=sedes,
        turnos=turnos,
        selected_sede=selected_sede,
        selected_turno=selected_turno,
        selected_date_input=selected_date.strftime("%Y-%m-%d"),
        selected_date_label=selected_date.strftime("%d/%m/%Y"),
        selected_sede_label=sede_label,
        selected_turno_label=turno_label,
        restaurant_name=RESTAURANT_NAME,
        closing_data=closing_data,
        dashboard_data=dashboard_data,
        is_admin=is_admin,
    )


@app.post("/api/cierre-caja")
def guardar_cierre_caja():
    payload = request.get_json(force=True) or {}
    raw_sede = payload.get("id_sede")
    raw_turno = payload.get("id_turno")
    raw_fecha = payload.get("fecha")
    if not raw_sede or not raw_turno or not raw_fecha:
        return jsonify({"error": "Sede, turno y fecha son obligatorios"}), 400

    try:
        id_sede = int(raw_sede)
    except (TypeError, ValueError):
        return jsonify({"error": "ID de sede inválido"}), 400

    try:
        fecha = datetime.fromisoformat(raw_fecha).date()
    except ValueError:
        return jsonify({"error": "Fecha inválida"}), 400

    monto_inicial = _safe_float(payload.get("monto_inicial"))
    pos = _safe_float(payload.get("pos"))
    yape = _safe_float(payload.get("yape"))
    plin = _safe_float(payload.get("plin"))
    efectivo = _safe_float(payload.get("efectivo"))
    venta_sistema = _safe_float(payload.get("venta_sistema"))
    observaciones = str(payload.get("observaciones", "")).strip()

    raw_gastos = payload.get("gastos")
    gastos = []
    total_gastos = 0.0
    if isinstance(raw_gastos, list):
        for gasto in raw_gastos:
            if not isinstance(gasto, dict):
                continue
            monto = _safe_float(gasto.get("monto"))
            descripcion = str(gasto.get("descripcion", "")).strip()
            gastos.append({"descripcion": descripcion, "monto": monto})
            total_gastos += monto

    total_ingresos = pos + yape + plin + efectivo
    total_actual = (monto_inicial + total_ingresos) - total_gastos
    diferencia = total_actual - venta_sistema

    with Session(engine) as db:
        cierre = db.scalar(
            select(CierreCaja).where(
                CierreCaja.ID_Sede == id_sede,
                CierreCaja.ID_Turno == raw_turno,
                CierreCaja.Fecha == fecha,
            )
        )
        if not cierre:
            cierre = CierreCaja(
                Fecha=fecha,
                ID_Sede=id_sede,
                ID_Turno=raw_turno,
            )
            db.add(cierre)
        cierre.Monto_Inicial = monto_inicial
        cierre.POS = pos
        cierre.Yape = yape
        cierre.Plin = plin
        cierre.Efectivo = efectivo
        cierre.Venta_Sistema = venta_sistema
        cierre.Observaciones = observaciones
        cierre.Total_Gastos = total_gastos
        cierre.Total_Actual = total_actual
        cierre.Diferencia = diferencia
        cierre.Gastos_Detalle = json.dumps(gastos, ensure_ascii=False)
        db.commit()
        db.refresh(cierre)

    response_payload = {
        "id": cierre.ID_Cierre,
        "fecha": fecha.isoformat(),
        "id_sede": cierre.ID_Sede,
        "id_turno": cierre.ID_Turno,
        "monto_inicial": cierre.Monto_Inicial,
        "pos": cierre.POS,
        "yape": cierre.Yape,
        "plin": cierre.Plin,
        "efectivo": cierre.Efectivo,
        "venta_sistema": cierre.Venta_Sistema,
        "total_gastos": cierre.Total_Gastos,
        "total_actual": cierre.Total_Actual,
        "diferencia": cierre.Diferencia,
        "observaciones": cierre.Observaciones,
        "gastos": gastos,
    }
    return jsonify({"ok": True, "cierre": response_payload})


@app.get("/api/checklist/items")
def checklist_items():
    if not _is_sede_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    id_sede = session.get("id_sede")
    id_area = session.get("id_area")
    id_turno = session.get("id_turno")
    ctx = {
        "id_sede": id_sede,
        "id_turno": id_turno,
        "id_area": id_area,
        "id_usuario": session.get("id_usuario"),
    }
    fecha_pedido = obtener_fecha_checklist()
    fecha_anterior = fecha_pedido - timedelta(days=1)
    with Session(engine) as db:
        pedido = get_or_create_pedido(db, ctx, fecha_pedido)
        detalles = {
            d.ID_Producto: d
            for d in db.scalars(
                select(DetallePedido).where(DetallePedido.ID_Pedido == pedido.ID_Pedido)
            ).all()
        }
        ocultos = set(
            db.scalars(
                select(ChecklistProductoOculto.ID_Producto).where(
                    and_(
                        ChecklistProductoOculto.ID_Sede == id_sede,
                        ChecklistProductoOculto.ID_Area == id_area,
                    )
                )
            ).all()
        )
        productos = db.scalars(
            select(Producto).where(
                and_(
                    Producto.Estado == "Activo",
                    Producto.ID_Area == id_area,
                )
            )
        ).all()
        result = []
        for p in productos:
            if p.ID_Producto in ocultos:
                continue
            inv = db.get(InventarioSede, (id_sede, p.ID_Producto))
            detalle = detalles.get(p.ID_Producto)
            detalle_estado = detalle.Estado_Sede if detalle else "Pendiente"
            unidad = db.get(Unidad, p.ID_Unidad) if p.ID_Unidad else None
            result.append(
                {
                    "id_producto": p.ID_Producto,
                    "nombre_producto": p.Nombre_Producto,
                    "subarea": p.Subarea or "Sin subárea",
                    "subarea_badge": subarea_badge(p.ID_Area, p.Subarea),
                    "stock_actual_sede": inv.Stock_Actual if inv else 0,
                    "detalle_id": detalle.ID_Detalle if detalle else None,
                    "pedido_id": pedido.ID_Pedido,
                    "detalle_estado": detalle_estado,
                    "cantidad_solicitada": detalle.Cantidad_Pedida if detalle else None,
                    "cantidad_entregada": detalle.Cantidad_Entregada if detalle else 0,
                    "check_almacen": bool(detalle.Check_Almacen) if detalle else False,
                    "unidad": unidad.Nombre_Unidad if unidad else None,
                    "area_id": p.ID_Area,
                }
            )
        def sort_key(item: dict[str, str | int | None]) -> tuple[int, int, str]:
            area_rank = AREA_ORDER.get(item.get("area_id"), len(AREA_ORDER))
            sub_map = SUBAREA_POSITION.get(item.get("area_id"), {})
            sub_rank = sub_map.get(normalize_subarea_name(item.get("subarea")), len(sub_map))
            return area_rank, sub_rank, item.get("nombre_producto", "") or ""
        result.sort(key=sort_key)
    return jsonify(result)


@app.get("/api/checklist/catalogo")
def checklist_catalogo():
    if not _is_sede_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    id_sede = session.get("id_sede")
    id_area = session.get("id_area")
    ctx = {
        "id_sede": id_sede,
        "id_turno": session.get("id_turno"),
        "id_area": id_area,
        "id_usuario": session.get("id_usuario"),
    }
    fecha_pedido = obtener_fecha_checklist()
    fecha_anterior = fecha_pedido - timedelta(days=1)
    with Session(engine) as db:
        productos = db.scalars(
            select(Producto).where(
                and_(
                    Producto.Estado == "Activo",
                    Producto.ID_Area == id_area,
                )
            )
        ).all()
        result = []
        historico_nombres = obtener_lista_historica(db, ctx, fecha_anterior)
        filtro_activo = bool(historico_nombres)
        historico_set = {nombre for nombre in historico_nombres}
        recent_products = set(session.get('recent_catalog_products') or [])
        for p in productos:
            if filtro_activo and p.Nombre_Producto not in historico_set and p.ID_Producto not in recent_products:
                continue
            inv = db.get(InventarioSede, (id_sede, p.ID_Producto))
            unidad = db.get(Unidad, p.ID_Unidad) if p.ID_Unidad else None
            result.append(
                {
                    "id_producto": p.ID_Producto,
                    "nombre_producto": p.Nombre_Producto,
                    "subarea": p.Subarea or "Sin subárea",
                    "subarea_badge": subarea_badge(p.ID_Area, p.Subarea),
                    "stock_actual_sede": inv.Stock_Actual if inv else 0,
                    "unidad": unidad.Nombre_Unidad if unidad else None,
                    "area_id": p.ID_Area,
                }
            )
        def sort_key(item: dict[str, str | int | None]) -> tuple[int, int, str]:
            area_rank = AREA_ORDER.get(item.get("area_id"), len(AREA_ORDER))
            sub_map = SUBAREA_POSITION.get(item.get("area_id"), {})
            sub_rank = sub_map.get(normalize_subarea_name(item.get("subarea")), len(sub_map))
            return area_rank, sub_rank, item.get("nombre_producto", "") or ""
        result.sort(key=sort_key)
    return jsonify(result)


@app.post("/api/pedidos/solicitar")
def solicitar_item():
    if not _is_sede_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    payload = request.get_json(force=True)
    id_producto = payload.get("id_producto")
    pedido_id_payload = payload.get("pedido_id")
    try:
        cantidad = float(payload.get("cantidad", 0))
    except (TypeError, ValueError):
        cantidad = 0

    if not id_producto:
        return jsonify({"error": "Producto requerido"}), 400
    if cantidad <= 0:
        return jsonify({"error": "La cantidad debe ser mayor que 0"}), 400

    ctx = {
        "id_sede": session.get("id_sede"),
        "id_turno": session.get("id_turno"),
        "id_area": session.get("id_area"),
        "id_usuario": session.get("id_usuario"),
    }

    fecha_pedido = obtener_fecha_checklist()
    with Session(engine) as db:
        if pedido_id_payload:
            pedido = db.get(ChecklistPedido, int(pedido_id_payload))
        else:
            pedido = get_or_create_pedido(db, ctx, fecha_pedido)

        if (
            not pedido
            or pedido.ID_Sede != ctx["id_sede"]
            or pedido.ID_Turno != ctx["id_turno"]
            or pedido.ID_Area != ctx["id_area"]
        ):
            return jsonify({"error": "Pedido no válido"}), 400
        sede = db.get(Sede, ctx["id_sede"])
        area = db.get(Area, ctx["id_area"])
        turno = db.get(Turno, ctx["id_turno"])
        db.execute(
            delete(ChecklistProductoOculto).where(
                and_(
                    ChecklistProductoOculto.ID_Sede == ctx["id_sede"],
                    ChecklistProductoOculto.ID_Area == ctx["id_area"],
                    ChecklistProductoOculto.ID_Producto == id_producto,
                )
            )
        )
        detalle = db.scalar(
            select(DetallePedido).where(
                and_(
                    DetallePedido.ID_Pedido == pedido.ID_Pedido,
                    DetallePedido.ID_Producto == id_producto,
                )
            )
        )

        if detalle:
            if detalle.Estado_Sede == "Enviado":
                return jsonify({"error": "El insumo ya fue enviado"}), 400
            detalle.Cantidad_Pedida = cantidad
            detalle.Estado_Sede = "Listo"
        else:
            detalle = DetallePedido(
                ID_Pedido=pedido.ID_Pedido,
                ID_Producto=id_producto,
                Cantidad_Pedida=cantidad,
                Estado_Sede="Listo",
            )
            db.add(detalle)
            pedido.Estado_General = "Pendiente"

        detalles_actuales = db.scalars(
            select(DetallePedido).where(DetallePedido.ID_Pedido == pedido.ID_Pedido)
        ).all()
        actualizar_historial_checklist(db, ctx, fecha_pedido, detalles_actuales)
        db.commit()
        estado_actual = pedido.Estado_General
        detalle_estado = detalle.Estado_Sede

    return jsonify({"ok": True, "estado": estado_actual, "detalle_estado": detalle_estado})


@app.delete("/api/pedidos/detalle")
def quitar_item():
    if not _is_sede_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    payload = request.get_json(force=True)
    pedido_id = payload.get("pedido_id")
    id_producto = payload.get("id_producto")

    if not pedido_id or not id_producto:
        return jsonify({"error": "Datos incompletos"}), 400

    ctx = {
        "id_sede": session.get("id_sede"),
        "id_turno": session.get("id_turno"),
        "id_area": session.get("id_area"),
    }
    fecha_pedido = obtener_fecha_checklist()

    with Session(engine) as db:
        pedido = db.get(ChecklistPedido, int(pedido_id))
        if (
            not pedido
            or pedido.ID_Sede != ctx["id_sede"]
            or pedido.ID_Turno != ctx["id_turno"]
            or pedido.ID_Area != ctx["id_area"]
        ):
            return jsonify({"error": "Pedido no válido"}), 400

        if pedido.Estado_General != "Pendiente":
            return (
                jsonify({"error": "El pedido ya fue enviado y no puede modificarse"}),
                400,
            )

        detalle = db.scalar(
            select(DetallePedido).where(
                and_(
                    DetallePedido.ID_Pedido == pedido.ID_Pedido,
                    DetallePedido.ID_Producto == id_producto,
                )
            )
        )
        if not detalle:
            return jsonify({"error": "Insumo no encontrado en el pedido"}), 404
        if detalle.Estado_Sede == "Enviado":
            return jsonify({"error": "No se puede quitar un insumo ya enviado"}), 400

        db.delete(detalle)
        existente = db.get(
            ChecklistProductoOculto,
            (ctx["id_sede"], ctx["id_area"], id_producto),
        )
        if not existente:
            db.add(
                ChecklistProductoOculto(
                    ID_Sede=ctx["id_sede"],
                    ID_Area=ctx["id_area"],
                    ID_Producto=id_producto,
                )
            )
        pedido.Estado_General = "Pendiente"
        detalles_actuales = db.scalars(
            select(DetallePedido).where(DetallePedido.ID_Pedido == pedido.ID_Pedido)
        ).all()
        actualizar_historial_checklist(db, ctx, fecha_pedido, detalles_actuales)
        db.commit()

    return jsonify({"ok": True, "estado": pedido.Estado_General})


@app.post("/api/checklist/productos")
def crear_producto_desde_checklist():
    if not _is_sede_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    payload = request.get_json(force=True)
    password = payload.get("password", "").strip()
    nombre = payload.get("nombre", "").strip()
    raw_subarea = payload.get("subarea")
    subarea = raw_subarea.strip() if isinstance(raw_subarea, str) else ""
    id_area = session.get("id_area")

    if not password:
        return jsonify({"error": "Contraseña requerida"}), 400

    if not nombre:
        return jsonify({"error": "Nombre obligatorio"}), 400
    if id_area not in SUBAREAS_CONFIG:
        return jsonify({"error": "Área inválida"}), 400

    subareas_validas = [item["nombre"] for item in SUBAREAS_CONFIG[id_area]]
    if not subareas_validas:
        return jsonify({"error": "El área no tiene subáreas configuradas"}), 400
    if not subarea:
        subarea = subareas_validas[0]
    elif subarea not in subareas_validas:
        return jsonify({"error": "Subárea inválida"}), 400

    id_categoria_payload = payload.get("id_categoria")
    id_unidad_payload = payload.get("id_unidad")

    with Session(engine) as db:
        if not verify_creation_password(db, password):
            return jsonify({"error": "Contraseña inválida"}), 403
        existe = db.scalar(
            select(Producto).where(
                and_(
                    Producto.Nombre_Producto == nombre,
                    Producto.ID_Area == id_area,
                    Producto.Subarea == subarea,
                )
            )
        )
        if existe:
            return jsonify({"error": "Ese producto ya existe en esa subárea"}), 400

        # permitir especificar categoria/unidad desde checklist
        categoria = None
        unidad = None
        if id_categoria_payload:
            categoria = db.get(Categoria, int(id_categoria_payload))
        if id_unidad_payload:
            unidad = db.get(Unidad, int(id_unidad_payload))

        if not categoria:
            categoria = db.scalar(select(Categoria).order_by(Categoria.ID_Categoria))
        if not unidad:
            unidad = db.scalar(select(Unidad).order_by(Unidad.ID_Unidad))
        if not categoria or not unidad:
            return jsonify({"error": "Faltan catálogos de categoría/unidad"}), 400

        id_producto = next_product_id(db)
        extra_sedes = []
        sede_actual = session.get("id_sede")
        if sede_actual:
            extra_sedes.append(sede_actual)

        crear_producto_en_inventario(
            db=db,
            id_producto=id_producto,
            nombre=nombre,
            id_area=id_area,
            subarea=subarea,
            id_categoria=categoria.ID_Categoria,
            id_unidad=unidad.ID_Unidad,
            stock_central=DEFAULT_CHECKLIST_STOCK,
            extra_sedes=extra_sedes or None,
        )
        db.commit()
    recent = session.get('recent_catalog_products') or []
    if id_producto not in recent:
        recent.append(id_producto)
    session['recent_catalog_products'] = recent

    return jsonify({"ok": True, "id_producto": id_producto})


@app.post("/api/pedidos/enviar")
def enviar_pedido():
    if not _is_sede_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    payload = request.get_json(force=True)
    pedido_id = int(payload.get("pedido_id"))
    ctx = {
        "id_sede": session.get("id_sede"),
        "id_turno": session.get("id_turno"),
        "id_area": session.get("id_area"),
    }

    with Session(engine) as db:
        pedido = db.get(ChecklistPedido, pedido_id)
        if (
            not pedido
            or pedido.ID_Sede != ctx["id_sede"]
            or pedido.ID_Turno != ctx["id_turno"]
            or pedido.ID_Area != ctx["id_area"]
        ):
            return jsonify({"error": "Pedido no válido"}), 400

        detalles_listos = db.scalars(
            select(DetallePedido).where(
                and_(
                    DetallePedido.ID_Pedido == pedido.ID_Pedido,
                    DetallePedido.Estado_Sede == "Listo",
                )
            )
        ).all()
        if not detalles_listos:
            return jsonify({"error": "No hay insumos listos para enviar"}), 400

        enviados_ids = []
        for detalle in detalles_listos:
            detalle.Estado_Sede = "Enviado"
            enviados_ids.append(detalle.ID_Producto)

        pedido.Estado_General = "Parcial"
        db.commit()
        estado_actual = pedido.Estado_General

    return jsonify({"ok": True, "estado": estado_actual, "enviados": enviados_ids})


@app.post("/api/pedidos/llegada")
def confirmar_llegada():
    if not _is_sede_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    payload = request.get_json(force=True)
    pedido_id = payload.get("pedido_id")
    raw_detalles = payload.get("detalle_ids", [])
    if not pedido_id:
        return jsonify({"error": "Pedido requerido"}), 400

    try:
        detalles_ids = [int(d) for d in (raw_detalles or []) if d is not None]
    except (TypeError, ValueError):
        return jsonify({"error": "IDs inválidos"}), 400

    if not detalles_ids:
        return jsonify({"error": "No se seleccionaron insumos"}), 400

    ctx = {
        "id_sede": session.get("id_sede"),
        "id_turno": session.get("id_turno"),
        "id_area": session.get("id_area"),
    }

    try:
        with Session(engine) as db:
            pedido = db.get(ChecklistPedido, int(pedido_id))
            if not pedido or not pedido_pertenece_al_contexto(pedido, ctx):
                return jsonify({"error": "Pedido no válido"}), 400

            detalles = db.scalars(
                select(DetallePedido).where(
                    and_(
                        DetallePedido.ID_Pedido == pedido.ID_Pedido,
                        DetallePedido.ID_Detalle.in_(detalles_ids),
                    )
                )
            ).all()

            actualizados = 0
            for detalle in detalles:
                if detalle.Estado_Sede == "Enviado":
                    detalle.Estado_Sede = "Recibido"
                    actualizados += 1

            if actualizados == 0:
                return (
                    jsonify({
                        "ok": True,
                        "updated": 0,
                        "estado": pedido.Estado_General,
                        "message": "No había insumos pendientes de llegada",
                    }),
                    200,
                )

            actualizar_estado_general_pedido(pedido)
            db.commit()

        return jsonify({"ok": True, "updated": actualizados, "estado": pedido.Estado_General})
    except Exception as exc:
        app.logger.exception("Error al confirmar llegada de insumos", exc_info=exc)
        return (
            jsonify({"error": "No se pudo confirmar la llegada. Intenta de nuevo."}),
            500,
        )


@app.post("/api/pedidos/confirmar-recibidos")
def confirmar_recibidos():
    if not _is_sede_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    payload = request.get_json(force=True)
    pedido_id = payload.get("pedido_id")
    if not pedido_id:
        return jsonify({"error": "Pedido requerido"}), 400

    ctx = {
        "id_sede": session.get("id_sede"),
        "id_turno": session.get("id_turno"),
        "id_area": session.get("id_area"),
    }

    with Session(engine) as db:
        pedido = db.get(ChecklistPedido, int(pedido_id))
        if not pedido or not pedido_pertenece_al_contexto(pedido, ctx):
            return jsonify({"error": "Pedido no válido"}), 400

        actualizados = 0
        for detalle in pedido.detalles:
            if detalle.Estado_Sede == "Enviado":
                detalle.Estado_Sede = "Recibido"
                actualizados += 1

        if actualizados == 0:
            return jsonify({
                "ok": True,
                "updated": 0,
                "estado": pedido.Estado_General,
                "message": "No había insumos pendientes de llegada",
            })

        actualizar_estado_general_pedido(pedido)
        db.commit()

    return jsonify({"ok": True, "updated": actualizados, "estado": pedido.Estado_General})


@app.get("/almacen")
def almacen():
    if not _is_almacen_or_admin_profile():
        return redirect(url_for("login"))

    raw_fecha = request.args.get("fecha")
    selected_date = None
    if raw_fecha:
        try:
            selected_date = datetime.fromisoformat(raw_fecha).date()
        except ValueError:
            selected_date = None
    if not selected_date:
        selected_date = datetime.now().date()
    selected_date_input = selected_date.strftime("%Y-%m-%d")
    selected_date_label = selected_date.strftime("%d/%m/%Y")
    movimientos_only = bool(session.get("movimientos_only"))

    is_admin = session.get("perfil") == "admin"
    current_sede = session.get("id_sede")
    current_turno = session.get("id_turno")
    with Session(engine) as db:
        categorias = db.scalars(select(Categoria).order_by(Categoria.Nombre_Categoria)).all()
        unidades = db.scalars(select(Unidad).order_by(Unidad.Nombre_Unidad)).all()
        areas = db.scalars(select(Area).order_by(Area.Nombre_Area)).all()
        productos_activos = db.scalars(select(Producto).where(Producto.Estado == 'Activo').order_by(Producto.Nombre_Producto)).all()
        almacen_products: list[dict[str, str | float]] = []
        for producto in productos_activos:
            unit = db.get(Unidad, producto.ID_Unidad) if producto.ID_Unidad else None
            inv = db.get(InventarioSede, (CENTRAL_SEDE_ID, producto.ID_Producto))
            almacen_products.append({
                'id': producto.ID_Producto,
                'nombre': producto.Nombre_Producto,
                'unidad': unit.Nombre_Unidad if unit else '',
                'stock': float(inv.Stock_Actual or 0) if inv else 0,
            })
    return render_template(
        "almacen.html",
        categorias=categorias,
        unidades=unidades,
        areas=areas,
        subareas_config=SUBAREAS_CONFIG,
        selected_date_input=selected_date_input,
        selected_date_label=selected_date_label,
        is_admin=is_admin,
        current_sede=current_sede,
        current_turno=current_turno,
        almacen_products=almacen_products,
        movimientos_only=movimientos_only,
    )


@app.route("/almacen/ajustes", methods=["GET", "POST"])
def almacen_ajustes():
    if not _is_almacen_or_admin_profile():
        return redirect(url_for("login"))

    message = None
    error = None
    message_type = "info"
    user_management_state = load_user_management_state()
    with Session(engine) as db:
        sedes = db.scalars(
            select(Sede).where(Sede.ID_Sede.in_([17, 20])).order_by(Sede.ID_Sede)
        ).all()
        areas = db.scalars(select(Area).order_by(Area.Nombre_Area)).all()
        turnos = db.scalars(select(Turno).order_by(Turno.Nombre_Turno)).all()
        admin_credentials = {
            "usuario": get_setting_value(db, ADMIN_USER_KEY, DEFAULT_ADMIN_CREDENTIALS["usuario"]),
            "password": get_setting_value(db, ADMIN_PASSWORD_KEY, DEFAULT_ADMIN_CREDENTIALS["password"]),
        }
        default_almacen_credentials = {
            "usuario": get_setting_value(db, ALMACEN_USER_KEY, DEFAULT_ALMACEN_CREDENTIALS["usuario"]),
            "password": get_setting_value(db, ALMACEN_PASSWORD_KEY, DEFAULT_ALMACEN_CREDENTIALS["password"]),
        }
    if request.method == "POST":
        action = request.form.get("form_action", "change_password")
        if action == "update_users":
            cocina_list, parse_error = _parse_user_updates("cocina", user_management_state["cocina"], request.form)
            if parse_error:
                error = parse_error
            else:
                caja_list, parse_error = _parse_user_updates("caja", user_management_state["caja"], request.form)
                if parse_error:
                    error = parse_error
                else:
                    cierre_list, parse_error = _parse_user_updates("cierre", user_management_state["cierre"], request.form)
                    if parse_error:
                        error = parse_error
                    else:
                        movimientos_list, parse_error = _parse_user_updates("movimientos", user_management_state["movimientos"], request.form)
                        if parse_error:
                            error = parse_error
                        else:
                            persist_user_group("cocina", cocina_list)
                            persist_user_group("caja", caja_list)
                            persist_user_group("cierre", cierre_list)
                            persist_user_group("movimientos", movimientos_list)
                            user_management_state = load_user_management_state()
                            message = "Credenciales actualizadas correctamente"
                            message_type = "success"
        elif action == "update_admin_credentials":
            admin_usuario = request.form.get("admin_usuario", "").strip()
            admin_password = request.form.get("admin_password", "")
            almacen_usuario = request.form.get("almacen_usuario", "").strip()
            almacen_password = request.form.get("almacen_password", "")
            if not admin_usuario or not admin_password or not almacen_usuario or not almacen_password:
                error = "Todos los campos son obligatorios"
            else:
                with Session(engine) as db:
                    set_setting_value(db, ADMIN_USER_KEY, admin_usuario)
                    set_setting_value(db, ADMIN_PASSWORD_KEY, admin_password)
                    set_setting_value(db, ALMACEN_USER_KEY, almacen_usuario)
                    set_setting_value(db, ALMACEN_PASSWORD_KEY, almacen_password)
                    db.commit()
                admin_credentials = {"usuario": admin_usuario, "password": admin_password}
                default_almacen_credentials = {"usuario": almacen_usuario, "password": almacen_password}
                message = "Credenciales críticas guardadas correctamente"
                message_type = "success"
        elif action == "update_admin_credentials":
            admin_usuario = request.form.get("admin_usuario", "").strip()
            admin_password = request.form.get("admin_password", "")
            almacen_usuario = request.form.get("almacen_usuario", "").strip()
            almacen_password = request.form.get("almacen_password", "")
            if not admin_usuario or not admin_password or not almacen_usuario or not almacen_password:
                error = "Todos los campos son obligatorios"
            else:
                with Session(engine) as db:
                    set_setting_value(db, ADMIN_USER_KEY, admin_usuario)
                    set_setting_value(db, ADMIN_PASSWORD_KEY, admin_password)
                    set_setting_value(db, ALMACEN_USER_KEY, almacen_usuario)
                    set_setting_value(db, ALMACEN_PASSWORD_KEY, almacen_password)
                    db.commit()
                admin_credentials = {"usuario": admin_usuario, "password": admin_password}
                default_almacen_credentials = {"usuario": almacen_usuario, "password": almacen_password}
                message = "Credenciales críticas guardadas correctamente"
                message_type = "success"
        else:
            current_password = request.form.get("current_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")
            if not current_password or not new_password or not confirm_password:
                error = "Todos los campos son obligatorios"
            elif new_password != confirm_password:
                error = "Las contraseñas no coinciden"
            else:
                with Session(engine) as db:
                    if not verify_creation_password(db, current_password):
                        error = "La contraseña actual no coincide"
                    else:
                        update_creation_password(db, new_password)
                        message = "Contraseña actualizada correctamente"
                        message_type = "success"

    return render_template(
        "almacen_ajustes.html",
        message=message,
        error=error,
        message_type=message_type,
        creation_password_hint=DEFAULT_CREATION_PASSWORD,
        user_management_cocina=user_management_state["cocina"],
        user_management_caja=user_management_state["caja"],
        user_management_cierre=user_management_state["cierre"],
        user_management_movimientos=user_management_state["movimientos"],
        sedes=sedes,
        areas=areas,
        turnos=turnos,
        admin_credentials=admin_credentials,
        default_almacen_credentials=default_almacen_credentials,
    )


@app.get("/api/almacen/pedidos")
def pedidos_almacen():
    if not _is_almacen_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    raw_fecha = request.args.get("fecha")
    fecha_obj = None
    if raw_fecha:
        try:
            fecha_obj = datetime.fromisoformat(raw_fecha).date()
        except ValueError:
            fecha_obj = None
    if not fecha_obj:
        fecha_obj = datetime.now().date()
    inicio = datetime.combine(fecha_obj, time.min)
    fin = datetime.combine(fecha_obj, time.max)

    with Session(engine) as db:
        stmt = select(ChecklistPedido).where(ChecklistPedido.Estado_General.in_(["Pendiente", "Parcial", "Enviado"]))
        stmt = stmt.where(and_(ChecklistPedido.Fecha >= inicio, ChecklistPedido.Fecha <= fin))
        pedidos = db.scalars(stmt).all()

        data = []
        for p in pedidos:
            sede = db.get(Sede, p.ID_Sede)
            turno = db.get(Turno, p.ID_Turno) if p.ID_Turno else None
            detalles = db.scalars(select(DetallePedido).where(DetallePedido.ID_Pedido == p.ID_Pedido)).all()
            det_data = []
            for d in detalles:
                producto = db.get(Producto, d.ID_Producto)
                stock_central = db.get(InventarioSede, (CENTRAL_SEDE_ID, d.ID_Producto))
                det_data.append(
                    {
                        "id_detalle": d.ID_Detalle,
                        "id_producto": d.ID_Producto,
                        "producto": producto.Nombre_Producto if producto else d.ID_Producto,
                        "area": producto.ID_Area if producto else None,
                        "subarea": producto.Subarea if producto else None,
                        "subarea_badge": subarea_badge(producto.ID_Area, producto.Subarea) if producto else "bg-secondary",
                        "cantidad_pedida": d.Cantidad_Pedida,
                        "check_almacen": d.Check_Almacen,
                        "cantidad_entregada": d.Cantidad_Entregada,
                        "estado_sede": d.Estado_Sede,
                        "stock_central": stock_central.Stock_Actual if stock_central else 0,
                    }
                )

            data.append(
                {
                    "id_pedido": p.ID_Pedido,
                    "sede": sede.Nombre_Sede if sede else str(p.ID_Sede),
                    "id_sede": p.ID_Sede,
                    "turno": turno.Nombre_Turno if turno else (p.ID_Turno or ""),
                    "turno_id": p.ID_Turno,
                    "estado": p.Estado_General,
                    "fecha": p.Fecha.isoformat() if p.Fecha else None,
                    "detalles": det_data,
                }
            )

        data.sort(key=lambda item: ((item.get("sede") or "").lower(), (item.get("turno") or "").lower()))

    return jsonify(data)


@app.get("/api/almacen/procesados")
def envios_procesados():
    if not _is_almacen_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    raw_fecha = request.args.get("fecha")
    fecha_obj = None
    if raw_fecha:
        try:
            fecha_obj = datetime.fromisoformat(raw_fecha).date()
        except ValueError:
            fecha_obj = None
    if not fecha_obj:
        fecha_obj = datetime.now().date()
    inicio = datetime.combine(fecha_obj, time.min)
    fin = datetime.combine(fecha_obj, time.max)

    with Session(engine) as db:
        stmt = select(ChecklistPedido).where(
            and_(
                ChecklistPedido.Fecha >= inicio,
                ChecklistPedido.Fecha <= fin,
            )
        )
        pedidos = db.scalars(stmt).all()
        data = []
        for p in pedidos:
            detalles = db.scalars(
                select(DetallePedido).where(
                    and_(
                        DetallePedido.ID_Pedido == p.ID_Pedido,
                        DetallePedido.Check_Almacen == 1,
                    )
                )
            ).all()
            if not detalles:
                continue
            sede = db.get(Sede, p.ID_Sede)
            turno = db.get(Turno, p.ID_Turno) if p.ID_Turno else None
            for d in detalles:
                producto = db.get(Producto, d.ID_Producto)
                data.append(
                    {
                        "sede": sede.Nombre_Sede if sede else str(p.ID_Sede),
                        "turno": turno.Nombre_Turno if turno else (p.ID_Turno or ""),
                        "producto": producto.Nombre_Producto if producto else d.ID_Producto,
                        "cantidad_pedida": d.Cantidad_Pedida,
                        "cantidad_entregada": d.Cantidad_Entregada,
                        "fecha": (p.Fecha.isoformat() if p.Fecha else None),
                    }
                )

    data.sort(key=lambda item: ((item.get("sede") or "").lower(), (item.get("turno") or "").lower(), item.get("producto") or ""))
    return jsonify(data)


@app.get("/api/almacen/movimientos")
def listar_movimientos_almacen():
    if not _is_almacen_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    raw_fecha = request.args.get("fecha")
    fecha_obj = None
    if raw_fecha:
        try:
            fecha_obj = datetime.fromisoformat(raw_fecha).date()
        except ValueError:
            fecha_obj = None
    if not fecha_obj:
        fecha_obj = datetime.now().date()
    inicio = datetime.combine(fecha_obj, time.min)
    fin = datetime.combine(fecha_obj, time.max)

    with Session(engine) as db:
        stmt = select(Movimiento).where(
            and_(
                Movimiento.Fecha >= inicio,
                Movimiento.Fecha <= fin,
            )
        ).order_by(Movimiento.Fecha.desc())
        movimientos = db.scalars(stmt).all()
        data = []
        for mov in movimientos:
            producto = db.get(Producto, mov.ID_Producto)
            unidad = db.get(Unidad, producto.ID_Unidad) if producto and producto.ID_Unidad else None
            data.append(
                {
                    "id_movimiento": mov.ID_Movimiento,
                    "tipo": mov.Tipo,
                    "producto": producto.Nombre_Producto if producto else mov.ID_Producto,
                    "unidad": unidad.Nombre_Unidad if unidad else '',
                    "cantidad": mov.Cantidad,
                    "motivo": mov.Motivo,
                    "usuario": mov.Usuario or '',
                    "fecha": mov.Fecha.isoformat() if mov.Fecha else None,
                }
            )

    return jsonify(data)


@app.get("/api/almacen/alertas/llegadas")
def alertas_llegadas():
    if not _is_almacen_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    raw_fecha = request.args.get("fecha")
    fecha_obj = None
    if raw_fecha:
        try:
            fecha_obj = datetime.fromisoformat(raw_fecha).date()
        except ValueError:
            fecha_obj = None
    if not fecha_obj:
        fecha_obj = datetime.now().date()
    inicio = datetime.combine(fecha_obj, time.min)
    fin = datetime.combine(fecha_obj, time.max)

    with Session(engine) as db:
        stmt = select(ChecklistPedido).where(
            and_(
                ChecklistPedido.Fecha >= inicio,
                ChecklistPedido.Fecha <= fin,
            )
        )
        pedidos = db.scalars(stmt).all()
        data = []
        for pedido in pedidos:
            sede = db.get(Sede, pedido.ID_Sede)
            turno = db.get(Turno, pedido.ID_Turno) if pedido.ID_Turno else None
            detalles = db.scalars(
                select(DetallePedido).where(
                    and_(
                        DetallePedido.ID_Pedido == pedido.ID_Pedido,
                        DetallePedido.Estado_Sede == "Enviado",
                        DetallePedido.Check_Almacen == 1,
                    )
                )
            ).all()
            if not detalles:
                continue
            for detalle in detalles:
                producto = db.get(Producto, detalle.ID_Producto)
                data.append(
                    {
                        "id_pedido": pedido.ID_Pedido,
                        "detalle_id": detalle.ID_Detalle,
                        "sede": sede.Nombre_Sede if sede else str(pedido.ID_Sede),
                        "turno": turno.Nombre_Turno if turno else (pedido.ID_Turno or ""),
                        "producto": producto.Nombre_Producto if producto else detalle.ID_Producto,
                        "cantidad_pedida": detalle.Cantidad_Pedida,
                            "cantidad_entregada": detalle.Cantidad_Entregada,
                            "check_almacen": bool(detalle.Check_Almacen),
                        "fecha": pedido.Fecha.isoformat() if pedido.Fecha else None,
                    }
                )

        data.sort(
            key=lambda item: (
                (item.get("sede") or "").lower(),
                (item.get("turno") or "").lower(),
                item.get("producto") or "",
            )
        )
    return jsonify(data)


@app.post("/api/almacen/procesar")
def procesar_envio():
    if not _is_almacen_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    payload = request.get_json(force=True)
    pedido_id = int(payload.get("pedido_id"))
    movimientos = payload.get("detalles", [])

    if not movimientos:
        return jsonify({"error": "No hay insumos seleccionados"}), 400

    with Session(engine) as db:
        pedido = db.get(ChecklistPedido, pedido_id)
        if not pedido:
            return jsonify({"error": "Pedido no encontrado"}), 404

        try:
            with db.begin_nested():
                procesados = 0
                for mov in movimientos:
                    id_detalle = int(mov["id_detalle"])
                    cantidad_entregada_raw = mov.get("cantidad_entregada")
                    detalle = db.get(DetallePedido, id_detalle)
                    if not detalle or detalle.ID_Pedido != pedido.ID_Pedido:
                        continue
                    if cantidad_entregada_raw is None:
                        cantidad_entregada_raw = detalle.Cantidad_Pedida
                    cantidad_entregada = float(cantidad_entregada_raw or 0)
                    if detalle.Check_Almacen:
                        raise ValueError("El insumo ya fue enviado")
                    if cantidad_entregada < 0:
                        raise ValueError("La cantidad entregada no puede ser negativa")
                    if cantidad_entregada == 0 and detalle.Cantidad_Pedida > 0:
                        # permitir registrar cero, pero avisar al usuario que no se entregó nada
                        pass

                    detalle.Check_Almacen = 1
                    detalle.Cantidad_Entregada = cantidad_entregada

                    if cantidad_entregada > 0:
                        inv_central = db.get(InventarioSede, (CENTRAL_SEDE_ID, detalle.ID_Producto))
                        inv_sede = db.get(InventarioSede, (pedido.ID_Sede, detalle.ID_Producto))

                        if not inv_central:
                            raise ValueError(f"Inventario central no encontrado para {detalle.ID_Producto}")
                        if not inv_sede:
                            inv_sede = InventarioSede(
                                ID_Sede=pedido.ID_Sede,
                                ID_Producto=detalle.ID_Producto,
                                Stock_Actual=0,
                                Punto_Minimo=0,
                            )
                            db.add(inv_sede)
                        if inv_central.Stock_Actual < cantidad_entregada:
                            raise ValueError(f"Stock insuficiente para {detalle.ID_Producto}")

                        inv_central.Stock_Actual -= cantidad_entregada
                        inv_sede.Stock_Actual += cantidad_entregada

                    procesados += 1

                if procesados == 0:
                    raise ValueError("No se pudo procesar ningún insumo")

                db.flush()
                if all(det.Check_Almacen for det in pedido.detalles):
                    pedido.Estado_General = "Recibido"
                else:
                    pedido.Estado_General = "Pendiente"

            db.commit()
            return jsonify({"ok": True, "estado": pedido.Estado_General})
        except ValueError as ex:
            db.rollback()
            return jsonify({"error": str(ex)}), 400


@app.post("/api/catalogo/categorias")
def crear_categoria():
    if not _is_almacen_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    nombre = request.get_json(force=True).get("nombre", "").strip()
    if not nombre:
        return jsonify({"error": "Nombre obligatorio"}), 400

    with Session(engine) as db:
        categoria = Categoria(Nombre_Categoria=nombre)
        db.add(categoria)
        db.commit()
        return jsonify({"ok": True, "id_categoria": categoria.ID_Categoria})


@app.post("/api/catalogo/unidades")
def crear_unidad():
    if not _is_almacen_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    nombre = request.get_json(force=True).get("nombre", "").strip()
    if not nombre:
        return jsonify({"error": "Nombre obligatorio"}), 400

    with Session(engine) as db:
        unidad = Unidad(Nombre_Unidad=nombre)
        db.add(unidad)
        db.commit()
        return jsonify({"ok": True, "id_unidad": unidad.ID_Unidad})


@app.get('/api/catalogo/categorias')
def listar_categorias():
    with Session(engine) as db:
        categorias = db.scalars(select(Categoria).order_by(Categoria.Nombre_Categoria)).all()
        return jsonify([{"id": c.ID_Categoria, "nombre": c.Nombre_Categoria} for c in categorias])


@app.get('/api/catalogo/unidades')
def listar_unidades():
    with Session(engine) as db:
        unidades = db.scalars(select(Unidad).order_by(Unidad.Nombre_Unidad)).all()
        return jsonify([{"id": u.ID_Unidad, "nombre": u.Nombre_Unidad} for u in unidades])


@app.post("/api/catalogo/productos")
def crear_producto():
    if not _is_almacen_or_admin_profile():
        return jsonify({"error": "No autorizado"}), 401

    payload = request.get_json(force=True)
    id_producto = payload.get("id_producto", "").strip()
    nombre = payload.get("nombre", "").strip()
    id_area = payload.get("id_area", "").strip()
    subarea = payload.get("subarea", "").strip()
    id_categoria = int(payload.get("id_categoria"))
    id_unidad = int(payload.get("id_unidad"))
    stock = float(payload.get("stock", 0) or 0)
    punto_minimo_raw = payload.get("punto_minimo")
    if punto_minimo_raw in (None, ""):
        punto_minimo = 5.0
    else:
        punto_minimo = float(punto_minimo_raw)
        if punto_minimo < 0:
            return jsonify({"error": "El punto mínimo no puede ser negativo"}), 400

    if not nombre:
        return jsonify({"error": "Nombre es obligatorio"}), 400
    if id_area not in SUBAREAS_CONFIG:
        return jsonify({"error": "Área inválida"}), 400

    subareas_validas = [item["nombre"] for item in SUBAREAS_CONFIG[id_area]]
    if subarea not in subareas_validas:
        return jsonify({"error": "Subárea inválida"}), 400

    with Session(engine) as db:
        # si no se proporciona ID, generarlo automáticamente
        if not id_producto:
            id_producto = next_product_id(db)

        existe = db.get(Producto, id_producto)
        if existe:
            return jsonify({"error": "El ID de producto ya existe"}), 400

        crear_producto_en_inventario(
            db=db,
            id_producto=id_producto,
            nombre=nombre,
            id_area=id_area,
            subarea=subarea,
            id_categoria=id_categoria,
            id_unidad=id_unidad,
            stock_central=stock,
            punto_minimo=punto_minimo,
        )

        db.commit()

    return jsonify({"ok": True})


@app.get('/api/catalogo/productos')
def listar_productos_catalogo():
    if not _is_almacen_or_admin_profile():
        return jsonify({'error': 'No autorizado'}), 401

    with Session(engine) as db:
        productos = db.scalars(
            select(Producto)
            .where(Producto.Estado == 'Activo')
            .order_by(Producto.Nombre_Producto)
        ).all()
        data = []
        for p in productos:
            inv = db.get(InventarioSede, (CENTRAL_SEDE_ID, p.ID_Producto))
            unidad = db.get(Unidad, p.ID_Unidad) if p.ID_Unidad else None
            categoria = db.get(Categoria, p.ID_Categoria) if p.ID_Categoria else None
            data.append(
                {
                    'id_producto': p.ID_Producto,
                    'nombre': p.Nombre_Producto,
                    'id_area': p.ID_Area,
                    'subarea': p.Subarea,
                    'id_categoria': p.ID_Categoria,
                    'id_unidad': p.ID_Unidad,
                    'categoria_nombre': categoria.Nombre_Categoria if categoria else None,
                    'unidad_nombre': unidad.Nombre_Unidad if unidad else None,
                    'stock_central': inv.Stock_Actual if inv else 0,
                    'punto_minimo': inv.Punto_Minimo if inv else 0,
                }
            )
    return jsonify(data)


@app.get('/api/inventario/productos/export')
def exportar_inventario():
    if not _is_almacen_or_admin_profile():
        return jsonify({'error': 'No autorizado'}), 401

    with Session(engine) as db:
        productos = db.scalars(select(Producto).order_by(Producto.Nombre_Producto)).all()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventario Central'
        ws.append(['ID', 'Nombre', 'Categoría', 'Subárea', 'Unidad', 'Punto mínimo', 'Stock Central', 'Estado'])
        for p in productos:
            inv = db.get(InventarioSede, (CENTRAL_SEDE_ID, p.ID_Producto))
            unidad = db.get(Unidad, p.ID_Unidad) if p.ID_Unidad else None
            categoria = db.get(Categoria, p.ID_Categoria) if p.ID_Categoria else None
            ws.append(
                [
                    p.ID_Producto,
                    p.Nombre_Producto,
                    categoria.Nombre_Categoria if categoria else '',
                    p.Subarea or '',
                    unidad.Nombre_Unidad if unidad else '',
                    inv.Punto_Minimo if inv else 0,
                    inv.Stock_Actual if inv else 0,
                    p.Estado,
                ]
            )
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name='inventario.xlsx',
            as_attachment=True,
        )


@app.post('/api/inventario/productos/import')
def importar_inventario():
    if not _is_almacen_or_admin_profile():
        return jsonify({'error': 'No autorizado'}), 401

    archivo = request.files.get('file')
    if not archivo:
        return jsonify({'error': 'Archivo requerido'}), 400

    try:
        workbook = load_workbook(archivo, data_only=True)
    except Exception as exc:
        return jsonify({'error': f'No se pudo leer el archivo: {exc}'}), 400

    sheet = workbook.active
    filas = list(sheet.iter_rows(values_only=True))
    if len(filas) < 2:
        return jsonify({'error': 'El archivo no contiene datos'}), 400

    header = [str(cell).strip().lower() if cell else '' for cell in filas[0]]
    columna_map = {name: idx for idx, name in enumerate(header) if name}

    def resolver_columna(*candidatos):
        for candidato in candidatos:
            if candidato in columna_map:
                return columna_map[candidato]
        return None

    id_idx = resolver_columna('id')
    nombre_idx = resolver_columna('nombre')
    categoria_idx = resolver_columna('categoría', 'categoria')
    subarea_idx = resolver_columna('subárea', 'subarea')
    unidad_idx = resolver_columna('unidad')
    stock_idx = resolver_columna('stock central')
    punto_idx = resolver_columna('punto mínimo', 'punto minimo')
    estado_idx = resolver_columna('estado')

    if id_idx is None or nombre_idx is None:
        return jsonify({'error': 'Las columnas "ID" y "Nombre" son obligatorias'}), 400

    def valor_seguro(fila, indice):
        if indice is None or indice >= len(fila):
            return None
        return fila[indice]

    def parsear_flotante(valor, defecto=0):
        if valor is None or (isinstance(valor, str) and not valor.strip()):
            return defecto
        try:
            return float(str(valor).replace(',', '.'))
        except ValueError:
            return defecto

    with Session(engine) as db:
        categorias = db.scalars(select(Categoria)).all()
        unidades = db.scalars(select(Unidad)).all()
        categorias_map = {c.Nombre_Categoria.strip().lower(): c.ID_Categoria for c in categorias if c.Nombre_Categoria}
        unidades_map = {u.Nombre_Unidad.strip().lower(): u.ID_Unidad for u in unidades if u.Nombre_Unidad}
        parsed_rows = []
        ids_vistos: set[str] = set()
        for fila_idx, fila in enumerate(filas[1:], start=2):
            raw_id = valor_seguro(fila, id_idx)
            prod_id = str(raw_id or '').strip()
            if not prod_id:
                continue
            if prod_id in ids_vistos:
                return jsonify({'error': f'El producto {prod_id} se repite en el archivo (fila {fila_idx})'}), 400
            ids_vistos.add(prod_id)

            nombre_raw = str(valor_seguro(fila, nombre_idx) or '').strip()
            if not nombre_raw:
                return jsonify({'error': f'El nombre es obligatorio en fila {fila_idx}'}), 400

            categoria_raw = str(valor_seguro(fila, categoria_idx) or '').strip() if categoria_idx is not None else ''
            if not categoria_raw:
                return jsonify({'error': f'La categoría es obligatoria en fila {fila_idx}'}), 400
            categoria_id = categorias_map.get(categoria_raw.lower())
            if not categoria_id:
                return jsonify({'error': f'Se encontró una categoría no existente: {categoria_raw} (fila {fila_idx})'}), 400

            subarea_raw = str(valor_seguro(fila, subarea_idx) or '').strip() if subarea_idx is not None else ''
            if not subarea_raw:
                return jsonify({'error': f'La subárea es obligatoria en fila {fila_idx}'}), 400
            area_id = SUBAREA_TO_AREA.get(subarea_raw.lower())
            if not area_id:
                return jsonify({'error': f'Se encontró una subárea no existente: {subarea_raw} (fila {fila_idx})'}), 400

            unidad_raw = str(valor_seguro(fila, unidad_idx) or '').strip() if unidad_idx is not None else ''
            if not unidad_raw:
                return jsonify({'error': f'La unidad es obligatoria en fila {fila_idx}'}), 400
            unidad_id = unidades_map.get(unidad_raw.lower())
            if not unidad_id:
                return jsonify({'error': f'Se encontró una unidad no existente: {unidad_raw} (fila {fila_idx})'}), 400

            stock = parsear_flotante(valor_seguro(fila, stock_idx), 0)
            punto = parsear_flotante(valor_seguro(fila, punto_idx), 5) if punto_idx is not None else 5
            estado_raw = str(valor_seguro(fila, estado_idx) or 'Activo').strip() if estado_idx is not None else 'Activo'
            estado_value = estado_raw or 'Activo'

            parsed_rows.append(
                {
                    'id': prod_id,
                    'nombre': nombre_raw,
                    'area_id': area_id,
                    'subarea': subarea_raw,
                    'categoria_id': categoria_id,
                    'unidad_id': unidad_id,
                    'stock': stock,
                    'punto_minimo': max(punto, 0),
                    'estado': estado_value,
                }
            )

        if not parsed_rows:
            return jsonify({'error': 'No se encontraron productos válidos en el archivo'}), 400

        created_count = 0
        updated_count = 0
        procesados = 0
        for registro in parsed_rows:
            producto = db.get(Producto, registro['id'])
            if producto:
                producto.Nombre_Producto = registro['nombre']
                producto.ID_Area = registro['area_id']
                producto.Subarea = registro['subarea']
                producto.ID_Categoria = registro['categoria_id']
                producto.ID_Unidad = registro['unidad_id']
                producto.Estado = registro['estado']
                updated_count += 1
            else:
                crear_producto_en_inventario(
                    db=db,
                    id_producto=registro['id'],
                    nombre=registro['nombre'],
                    id_area=registro['area_id'],
                    subarea=registro['subarea'],
                    id_categoria=registro['categoria_id'],
                    id_unidad=registro['unidad_id'],
                    stock_central=registro['stock'],
                    punto_minimo=registro['punto_minimo'],
                )
                producto = db.get(Producto, registro['id'])
                if producto:
                    producto.Estado = registro['estado']
                created_count += 1

            inv = db.get(InventarioSede, (CENTRAL_SEDE_ID, registro['id']))
            if not inv:
                inv = InventarioSede(
                    ID_Sede=CENTRAL_SEDE_ID,
                    ID_Producto=registro['id'],
                    Stock_Actual=0,
                    Punto_Minimo=5,
                )
                db.add(inv)
            inv.Stock_Actual = registro['stock']
            inv.Punto_Minimo = registro['punto_minimo']
            procesados += 1

        activos_importados = [registro['id'] for registro in parsed_rows]
        deactivated_count = 0
        if activos_importados:
            deactivated_count = db.query(Producto).filter(
                Producto.Estado == 'Activo',
                ~Producto.ID_Producto.in_(activos_importados)
            ).update({'Estado': 'Inactivo'}, synchronize_session=False)

        db.commit()

    mensaje = (
        f'Se crearon {created_count}, se actualizaron {updated_count}, se inactivaron {deactivated_count} productos.'
        if procesados
        else 'No se pudo importar ningún producto'
    )
    detalle = {
        'ok': True,
        'message': mensaje,
        'created': created_count,
        'updated': updated_count,
        'deactivated': deactivated_count,
        'processed': procesados,
    }
    return jsonify(detalle)


@app.post('/api/almacen/movimiento')
def registrar_movimiento():
    if not _is_almacen_or_admin_profile():
        return jsonify({'error': 'No autorizado'}), 401

    payload = request.get_json(force=True) or {}
    tipo = str(payload.get('tipo', '')).strip().lower()
    id_producto = str(payload.get('id_producto', '')).strip()
    motivo = str(payload.get('motivo', '')).strip()
    cantidad = _safe_float(payload.get('cantidad'))
    if tipo not in ('entrada', 'salida'):
        return jsonify({'error': 'Tipo inválido'}), 400
    if not id_producto:
        return jsonify({'error': 'Producto requerido'}), 400
    if cantidad <= 0:
        return jsonify({'error': 'Cantidad debe ser mayor a cero'}), 400
    if not motivo:
        return jsonify({'error': 'Motivo es obligatorio'}), 400

    with Session(engine) as db:
        producto = db.get(Producto, id_producto)
        if not producto:
            return jsonify({'error': 'Producto no encontrado'}), 404
        inventario = db.get(InventarioSede, (CENTRAL_SEDE_ID, id_producto))
        if not inventario:
            inventario = InventarioSede(ID_Sede=CENTRAL_SEDE_ID, ID_Producto=id_producto, Stock_Actual=0, Punto_Minimo=0)
            db.add(inventario)
            db.flush()
        stock_actual = float(inventario.Stock_Actual or 0)
        if tipo == 'entrada':
            inventario.Stock_Actual = stock_actual + cantidad
        else:
            if cantidad > stock_actual:
                return jsonify({'error': 'Stock insuficiente'}), 400
            inventario.Stock_Actual = stock_actual - cantidad
        movimiento = Movimiento(
            Tipo=tipo,
            ID_Producto=id_producto,
            Cantidad=cantidad,
            Motivo=motivo,
            Usuario=session.get('id_usuario') or session.get('perfil') or 'admin',
        )
        db.add(movimiento)
        db.commit()
        nuevo_stock = float(inventario.Stock_Actual or 0)
        unidad = db.get(Unidad, producto.ID_Unidad) if producto.ID_Unidad else None

    mensaje = 'Entrada registrada correctamente' if tipo == 'entrada' else 'Salida registrada correctamente'
    return jsonify({
        'ok': True,
        'tipo': tipo,
        'nuevo_stock': nuevo_stock,
        'mensaje': mensaje,
        'producto_nombre': producto.Nombre_Producto,
        'unidad': unidad.Nombre_Unidad if unidad else '',
    })


@app.get('/api/catalogo/productos/<id_producto>')
def obtener_producto(id_producto: str):
    if not _is_almacen_or_admin_profile():
        return jsonify({'error': 'No autorizado'}), 401

    with Session(engine) as db:
        p = db.get(Producto, id_producto)
        if not p:
            return jsonify({'error': 'Producto no encontrado'}), 404
        inv = db.get(InventarioSede, (CENTRAL_SEDE_ID, id_producto))
        unidad = db.get(Unidad, p.ID_Unidad) if p.ID_Unidad else None
        categoria = db.get(Categoria, p.ID_Categoria) if p.ID_Categoria else None
        return jsonify({
            'id_producto': p.ID_Producto,
            'nombre': p.Nombre_Producto,
            'id_area': p.ID_Area,
            'subarea': p.Subarea,
            'id_categoria': p.ID_Categoria,
            'id_unidad': p.ID_Unidad,
            'categoria_nombre': categoria.Nombre_Categoria if categoria else None,
            'unidad_nombre': unidad.Nombre_Unidad if unidad else None,
            'stock_central': inv.Stock_Actual if inv else 0,
            'punto_minimo': inv.Punto_Minimo if inv else 0,
        })


@app.delete('/api/catalogo/productos/<id_producto>')
def borrar_producto(id_producto: str):
    if not _is_almacen_or_admin_profile():
        return jsonify({'error': 'No autorizado'}), 401

    with Session(engine) as db:
        p = db.get(Producto, id_producto)
        if not p:
            return jsonify({'error': 'Producto no encontrado'}), 404
        # eliminar inventarios asociados
        db.query(InventarioSede).filter(InventarioSede.ID_Producto == id_producto).delete()
        db.delete(p)
        db.commit()
    return jsonify({'ok': True})


@app.put('/api/catalogo/productos/<id_producto>')
def actualizar_producto(id_producto: str):
    if not _is_almacen_or_admin_profile():
        return jsonify({'error': 'No autorizado'}), 401

    payload = request.get_json(force=True)
    nombre = payload.get('nombre')
    id_area = payload.get('id_area')
    subarea = payload.get('subarea')
    id_categoria = payload.get('id_categoria')
    id_unidad = payload.get('id_unidad')
    stock = payload.get('stock')
    punto_minimo = payload.get('punto_minimo')

    with Session(engine) as db:
        p = db.get(Producto, id_producto)
        if not p:
            return jsonify({'error': 'Producto no encontrado'}), 404

        if nombre is not None:
            p.Nombre_Producto = str(nombre).strip()
        if id_area is not None:
            p.ID_Area = id_area
        if subarea is not None:
            p.Subarea = subarea
        if id_categoria is not None:
            p.ID_Categoria = int(id_categoria)
        if id_unidad is not None:
            p.ID_Unidad = int(id_unidad)

        stock_value = None
        if stock is not None:
            stock_value = float(stock)

        punto_minimo_value = None
        if punto_minimo is not None:
            punto_minimo_value = float(punto_minimo)
            if punto_minimo_value < 0:
                return jsonify({'error': 'El punto mínimo no puede ser negativo'}), 400

        if stock is not None or punto_minimo is not None:
            inv = db.get(InventarioSede, (CENTRAL_SEDE_ID, id_producto))
            if not inv:
                inv = InventarioSede(
                    ID_Sede=CENTRAL_SEDE_ID,
                    ID_Producto=id_producto,
                    Stock_Actual=stock_value if stock_value is not None else 0,
                    Punto_Minimo=punto_minimo_value if punto_minimo_value is not None else 5,
                )
                db.add(inv)
            else:
                if stock_value is not None:
                    inv.Stock_Actual = stock_value
                if punto_minimo_value is not None:
                    inv.Punto_Minimo = punto_minimo_value

        db.commit()
        return jsonify({'ok': True})


@app.get("/logout")
def logout():
    _release_session_slot()
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
